from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook


FILES = {
    "Mirit Great American": Path(r"C:\Users\Sales\Downloads\Mirrit  (2).xlsx"),
    "Wilson Bros Van Lines": Path(r"C:\Users\Sales\Downloads\WIlson bros  (1).xlsx"),
    "Movers 95": Path(r"C:\Users\Sales\Downloads\Movers-95.xlsx"),
    "Gorilla Haulers": Path(r"C:\Users\Sales\Downloads\_Goruilla .xlsx"),
    "Top Tier Van Lines": Path(r"C:\Users\Sales\Downloads\Top tier pricing .xlsx"),
}

# company, sheet, header row, first rate row, min col, destination col,
# first/last band columns, service name/rate/comment columns, service first row
LAYOUTS = [
    ("Mirit Great American", "Sheet1", 10, 11, 1, 2, 3, 7, 9, 10, 11, 15),
    ("Wilson Bros Van Lines", "South East ", 10, 11, 1, 2, 3, 7, 9, 10, 11, 16),
    ("Wilson Bros Van Lines", "North east ", 10, 11, 1, 2, 3, 7, 9, 10, 11, 15),
    ("Wilson Bros Van Lines", "Midwest", 12, 13, 1, 2, 3, 11, 13, 14, 15, 14),
    ("Wilson Bros Van Lines", "CO ", 2, 3, 1, 2, 3, 6, 8, 9, 10, 11),
    ("Wilson Bros Van Lines", "northwest ", 7, 8, None, 1, 2, 4, 5, 6, 7, 19),
    ("Wilson Bros Van Lines", "Ca, NV, NM, AZ ", 6, 7, None, 1, 2, 4, 5, 6, 7, 10),
    ("Movers 95", "MD , VA , DC , DE", 15, 16, 2, 3, 4, 5, 11, 12, 13, 19),
    ("Gorilla Haulers", "East", 15, 16, 2, 3, 4, 5, 11, 12, 13, 19),
    ("Gorilla Haulers", "Midwest", 13, 14, 2, 3, 4, 5, 7, 8, 9, 17),
    ("Top Tier Van Lines", "East", 11, 12, 2, 3, 4, 5, 6, 7, None, 12),
    ("Top Tier Van Lines", "Midwest", 13, 14, 2, 3, 4, 5, 7, 8, 9, 17),
]


def text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:.10f}".rstrip("0").rstrip(".")
    return str(value).strip()


def numeric(value):
    if isinstance(value, (int, float)):
        return round(float(value), 4)
    try:
        return round(float(str(value).replace("$", "").replace(",", "").strip()), 4)
    except (TypeError, ValueError):
        return None


def band_bounds(label: str):
    normalized = label.lower().replace(",", "").replace("cf", "").strip()
    numbers = [int(value) for value in re.findall(r"\d+", normalized)]
    if not numbers:
        return None, None
    if "up" in normalized or "&" in normalized:
        return numbers[0], None
    if len(numbers) >= 2:
        return numbers[0], numbers[1]
    return numbers[0], numbers[0]


def main():
    books = {
        company: (
            load_workbook(path, data_only=False),
            load_workbook(path, data_only=True),
        )
        for company, path in FILES.items()
    }
    plans = []
    for sort_order, layout in enumerate(LAYOUTS):
        (
            company,
            sheet_name,
            header_row,
            first_rate_row,
            min_col,
            destination_col,
            first_band_col,
            last_band_col,
            service_name_col,
            service_rate_col,
            service_comment_col,
            first_service_row,
        ) = layout
        formula_sheet = books[company][0][sheet_name]
        value_sheet = books[company][1][sheet_name]
        source_path = FILES[company]
        band_labels = [
            text(value_sheet.cell(header_row, column).value)
            or text(formula_sheet.cell(header_row, column).value)
            for column in range(first_band_col, last_band_col + 1)
        ]
        rates = []
        empty_streak = 0
        region_group = ""
        for row in range(first_rate_row, formula_sheet.max_row + 1):
            destination = text(value_sheet.cell(row, destination_col).value)
            if not destination:
                empty_streak += 1
                if empty_streak >= 4:
                    break
                continue
            empty_streak = 0
            maybe_group = text(value_sheet.cell(row, max(1, destination_col - 2)).value)
            if maybe_group:
                region_group = maybe_group
            minimum_value = (
                value_sheet.cell(row, min_col).value if min_col is not None else None
            )
            minimum_text = text(
                formula_sheet.cell(row, min_col).value if min_col is not None else None
            )
            for offset, column in enumerate(range(first_band_col, last_band_col + 1)):
                raw = formula_sheet.cell(row, column).value
                resolved = value_sheet.cell(row, column).value
                if raw in (None, ""):
                    continue
                band_label = band_labels[offset] or f"Band {offset + 1}"
                lower, upper = band_bounds(band_label)
                rates.append(
                    {
                        "destination": destination,
                        "destination_group": region_group,
                        "minimum_price": numeric(minimum_value),
                        "minimum_text": minimum_text,
                        "band_label": band_label,
                        "cubic_feet_min": lower,
                        "cubic_feet_max": upper,
                        "rate": numeric(resolved),
                        "rate_text": text(raw),
                        "sort_order": len(rates),
                    }
                )

        services = []
        if service_name_col:
            empty_streak = 0
            for row in range(first_service_row, formula_sheet.max_row + 1):
                name = text(formula_sheet.cell(row, service_name_col).value)
                rate_text = text(formula_sheet.cell(row, service_rate_col).value)
                comments = (
                    text(formula_sheet.cell(row, service_comment_col).value)
                    if service_comment_col
                    else ""
                )
                if not any((name, rate_text, comments)):
                    empty_streak += 1
                    if empty_streak >= 6:
                        break
                    continue
                empty_streak = 0
                if name.lower() in {"item", "additional services"}:
                    continue
                services.append(
                    {
                        "name": name or "Additional note",
                        "rate_text": rate_text,
                        "comments": comments,
                        "sort_order": len(services),
                    }
                )

        # Preserve every top-of-sheet instruction/exception as a rule. This is
        # intentionally textual because several workbooks contain judgment calls.
        rules = []
        seen = set()
        for row in range(1, first_rate_row):
            parts = []
            for column in range(1, formula_sheet.max_column + 1):
                value = text(formula_sheet.cell(row, column).value)
                if value and value not in parts:
                    parts.append(value)
            value = " — ".join(parts)
            lower = value.lower()
            if not value or value in seen:
                continue
            if any(
                marker in lower
                for marker in (
                    "pick up", "fuel", "important", "rate", "exception",
                    "move", "mile", "toll", "minimum", "destination",
                    "deposit", "season", "discount", "ask",
                )
            ):
                seen.add(value)
                rules.append(
                    {
                        "category": (
                            "exception"
                            if any(word in lower for word in ("if ", "exception", "ask", "mile", "toll"))
                            else "general"
                        ),
                        "title": "Exception" if "if " in lower or "mile" in lower else "Pricing rule",
                        "description": value,
                        "sort_order": len(rules),
                    }
                )

        pickup_regions = ""
        for rule in rules:
            if "pick up from" in rule["description"].lower():
                pickup_regions = rule["description"]
                break
        if not pickup_regions:
            candidates = []
            for row in range(1, first_rate_row):
                for column in range(1, formula_sheet.max_column + 1):
                    value = text(formula_sheet.cell(row, column).value)
                    state_count = len(re.findall(r"\b[A-Z]{2}\b", value.upper()))
                    if state_count >= 3:
                        candidates.append((state_count, value))
            pickup_regions = max(candidates, key=lambda item: (item[0], len(item[1])), default=(0, ""))[1]
        if not pickup_regions and sheet_name.strip().upper() == "CO":
            pickup_regions = "Denver / Colorado pickups"
        fuel_match = re.search(
            r"fuel\s*(\d+(?:\.\d+)?)\s*%",
            " ".join(rule["description"] for rule in rules),
            re.IGNORECASE,
        )
        plans.append(
            {
                "source_key": f"{source_path.name}::{sheet_name}".lower(),
                "company_name": company,
                "name": sheet_name.strip() or "Default",
                "source_file": source_path.name,
                "source_sheet": sheet_name,
                "pickup_regions": pickup_regions,
                "fuel_percent": float(fuel_match.group(1)) if fuel_match else None,
                "active": True,
                "sort_order": sort_order,
                "rules": rules,
                "rates": rates,
                "services": services,
            }
        )
    output = Path("backend/pricing_seed.json")
    output.write_text(json.dumps(plans, indent=2), encoding="utf-8")
    print(
        f"Wrote {len(plans)} plans, {sum(len(p['rates']) for p in plans)} rates, "
        f"{sum(len(p['rules']) for p in plans)} rules, and "
        f"{sum(len(p['services']) for p in plans)} services to {output}"
    )


if __name__ == "__main__":
    main()

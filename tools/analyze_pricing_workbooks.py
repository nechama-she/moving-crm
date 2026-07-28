from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def useful_rows(sheet):
    rows = []
    for row in sheet.iter_rows():
        values = [cell.value for cell in row]
        if any(value not in (None, "") for value in values):
            last = max(index for index, value in enumerate(values) if value not in (None, ""))
            rows.append(
                {
                    "row": row[0].row,
                    "values": values[: last + 1],
                    "formulas": {
                        cell.coordinate: cell.value
                        for cell in row[: last + 1]
                        if cell.data_type == "f"
                    },
                    "comments": {
                        cell.coordinate: cell.comment.text
                        for cell in row[: last + 1]
                        if cell.comment
                    },
                }
            )
    return rows


def workbook_summary(path: Path):
    workbook = load_workbook(path, data_only=False, read_only=False)
    return {
        "file": str(path),
        "sheets": [
            {
                "title": sheet.title,
                "state": sheet.sheet_state,
                "dimensions": sheet.calculate_dimension(),
                "merged_cells": [str(cell_range) for cell_range in sheet.merged_cells.ranges],
                "hidden_rows": [
                    index for index, dimension in sheet.row_dimensions.items() if dimension.hidden
                ],
                "hidden_columns": [
                    get_column_letter(index)
                    for index, dimension in sheet.column_dimensions.items()
                    if dimension.hidden and isinstance(index, int)
                ]
                + [
                    index
                    for index, dimension in sheet.column_dimensions.items()
                    if dimension.hidden and isinstance(index, str)
                ],
                "data_validations": [
                    {
                        "ranges": str(validation.sqref),
                        "type": validation.type,
                        "formula1": validation.formula1,
                        "formula2": validation.formula2,
                    }
                    for validation in sheet.data_validations.dataValidation
                ],
                "rows": useful_rows(sheet),
            }
            for sheet in workbook.worksheets
        ],
    }


if __name__ == "__main__":
    summaries = [
        workbook_summary(Path(argument))
        for argument in sys.argv[1:]
        if argument != "--text"
    ]
    if "--text" in sys.argv:
        for summary in summaries:
            print(f"\n### {summary['file']}")
            for sheet in summary["sheets"]:
                print(f"\n## {sheet['title']}")
                for row in sheet["rows"]:
                    values = [
                        str(value)[:120] if value is not None else ""
                        for value in row["values"]
                    ]
                    print(row["row"], " | ".join(values))
    else:
        print(json.dumps(summaries, indent=2, default=str))
